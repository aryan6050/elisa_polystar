import openpyxl
import requests
import yaml
import logging
import os

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(script_dir)

config_path = os.path.join(parent_dir, 'config.yaml')
log_path = os.path.join(parent_dir, 'app.log')

logging.basicConfig(filename=log_path, level=logging.INFO)

def read_ilm_policies_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        ilm_policies = {}
        for row in range(2, sheet.max_row + 1):
            policy_name = sheet.cell(row=row, column=1).value
            print("policy name is ",policy_name)
            policy_hot_max_age = sheet.cell(row=row, column=2).value
            print("hot max age is ", policy_hot_max_age)
            policy_hot_max_size = sheet.cell(row=row, column=3).value
            print("hot max size is ", policy_hot_max_size)
            policy_warm_min_age = sheet.cell(row=row, column=4).value
            print("warm min ahe is ", policy_warm_min_age)
            policy_cold_min_age = sheet.cell(row=row, column=5).value
            print("cold min age is ", policy_cold_min_age)
            policy_delete_min_age = sheet.cell(row=row, column=6).value
            print("delete min age is ", policy_delete_min_age)

            required_fields = {
                "policy": {
                    "phases": {
                        "hot": {
                            "actions": {
                                "rollover": {
                                    # "max_age": (f"{policy_hot_max_age}"),
                                    # "max_size": (f"{policy_hot_max_size}")
                                }
                            }
                        },
                        "warm": {
                            # "min_age": (f"{policy_warm_min_age}"),
                            "actions": {
                                "allocate": {
                                    "require": {
                                        "box_type": "warm"
                                    }
                                }
                            }
                        },
                        "cold": {
                            # "min_age": (f"{policy_cold_min_age}"),
                            "actions": {
                                "allocate": {
                                    "require": {
                                        "box_type": "cold"
                                    }
                                }
                            }
                        },
                        "delete": {
                            # "min_age": (f"{policy_delete_min_age}"),
                            "actions": {
                                "delete": {}
                            }
                        }
                    }
                }
            }

            if policy_hot_max_age is None and policy_hot_max_size is None:
                del required_fields["policy"]["phases"]["hot"]
            else:
                if policy_hot_max_age is not None:
                    required_fields["policy"]["phases"]["hot"]["actions"]["rollover"]["max_age"] = (f"{policy_hot_max_age}")
                if policy_hot_max_size is not None:
                    required_fields["policy"]["phases"]["hot"]["actions"]["rollover"]["max_size"] = (f"{policy_hot_max_size}")
            
            if policy_warm_min_age is None:
                del required_fields["policy"]["phases"]["warm"]
            else:
                required_fields["policy"]["phases"]["warm"]["min_age"] = (f"{policy_hot_max_age}")
            
            if policy_cold_min_age is None:
                del required_fields["policy"]["phases"]["cold"]
            else:
                required_fields["policy"]["phases"]["cold"]["min_age"] = (f"{policy_cold_min_age}")

            if policy_delete_min_age is None:
                del required_fields["policy"]["phases"]["delete"]
            else:
                required_fields["policy"]["phases"]["delete"]["min_age"] = (f"{policy_delete_min_age}")
            
            
            ilm_policies[policy_name] = required_fields

            print("the ilm policy json here is ",ilm_policies)

        return ilm_policies
    except Exception as e:
        logging.error("Error:", e)
        return None

# Function to update ILM policies in Elasticsearch
def update_ilm_policies_in_elasticsearch(config):
    try:
        es_host = config["elasticsearch_attributes"]["host"]
        es_port = config["elasticsearch_attributes"]["port"]
        es_scheme = config["elasticsearch_attributes"]["scheme"]
        username = config["elasticsearch_attributes"]["username"]
        password = config["elasticsearch_attributes"]["password"]
        filename = os.path.join(parent_dir, 'update.xlsx')

        ilm_policies = read_ilm_policies_from_excel(filename)

        if ilm_policies:
            for policy_name, policy_body in ilm_policies.items():
                url = f"{es_scheme}://{es_host}:{es_port}/_ilm/policy/{policy_name}"
                headers = {'Content-Type': 'application/json'}
                response = requests.put(url, auth=(username, password), headers=headers, json=policy_body)
                if response.status_code == 200:
                    logging.info(f"ILM policy '{policy_name}' updated successfully.")
                else:
                    logging.error(f"Failed to update ILM policy '{policy_name}'. Status code: {response.status_code}")
                    logging.error(f"Response: {response.text}")
    except Exception as e:
        logging.error(f"Error: {e}")

if __name__ == '__main__':
    with open(config_path) as config_file:
        config = yaml.load(config_file, Loader=yaml.SafeLoader)
    update_ilm_policies_in_elasticsearch(config)
