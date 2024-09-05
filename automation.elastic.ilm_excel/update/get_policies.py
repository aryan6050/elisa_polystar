import openpyxl
import requests
import yaml
import logging
import os

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(script_dir)

log_path = os.path.join(parent_dir, 'app.log')

logging.basicConfig(filename=log_path, level=logging.INFO)

# Function to fetch all ILM policies from Elasticsearch
def fetch_ilm_policies(config):
    try:
        es_host = config["elasticsearch_attributes"]["host"]
        es_port = config["elasticsearch_attributes"]["port"]
        es_scheme = config["elasticsearch_attributes"]["scheme"]
        username = config["elasticsearch_attributes"]["username"]
        password = config["elasticsearch_attributes"]["password"]

        url = f"{es_scheme}://{es_host}:{es_port}/_ilm/policy"
        response = requests.get(url, auth=(username, password))
        if response.status_code == 200:
            policies = response.json()
            return policies
        else:
            logging.error(f"Failed to fetch ILM policies. Status code: {response.status_code}")
            return None
    except Exception as e:
        logging.error(f"Error: {e}")
        return None

# Function to store ILM policies in an Excel file
def store_ilm_policies_in_excel(ilm_policies, filename):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "ILM Policies"
        headers = ["policy_name", "policy_hot_max_age", "policy_hot_max_size", "policy_warm_min_age", "policy_cold_min_age", "policy_delete_min_age"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        
        for row, (policy_name, policy_data) in enumerate(ilm_policies.items(), start=2):
            sheet.cell(row=row, column=1, value=policy_name)
            if 'policy' in policy_data:
                policy_body = policy_data['policy']
                if 'phases' in policy_body and 'hot' in policy_body['phases']:
                    hot_phase = policy_body['phases']['hot']
                    sheet.cell(row=row, column=2, value=hot_phase.get('actions', {}).get('rollover', {}).get('max_age', ''))
                    sheet.cell(row=row, column=3, value=hot_phase.get('actions', {}).get('rollover', {}).get('max_size', ''))
                if 'phases' in policy_body and 'warm' in policy_body['phases']:
                    warm_phase = policy_body['phases']['warm']
                    sheet.cell(row=row, column=4, value=warm_phase.get('min_age', ''))
                if 'phases' in policy_body and 'cold' in policy_body['phases']:
                    cold_phase = policy_body['phases']['cold']
                    sheet.cell(row=row, column=5, value=cold_phase.get('min_age', ''))
                if 'phases' in policy_body and 'delete' in policy_body['phases']:
                    delete_phase = policy_body['phases']['delete']
                    sheet.cell(row=row, column=6, value=delete_phase.get('min_age', ''))
        
        wb.save(filename)
        logging.info(f"ILM policies stored in {filename} successfully.")
    except Exception as e:
        logging.error(f"Error: {e}")

if __name__ == '__main__':
    with open(os.path.join(parent_dir, 'config.yaml')) as config_file:
        config = yaml.load(config_file, Loader=yaml.SafeLoader)
    # Fetch ILM policies
    ilm_policies = fetch_ilm_policies(config)

    if ilm_policies:
        # Store ILM policies in Excel
       store_ilm_policies_in_excel(ilm_policies, os.path.join(parent_dir, 'update.xlsx'))
